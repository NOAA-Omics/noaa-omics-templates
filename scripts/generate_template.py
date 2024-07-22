import os
import re
import time
import sys
import gspread
import gspread_formatting as gf
from gspread_formatting import *
from gspread.utils import rowcol_to_a1
import pandas as pd
from google.oauth2.service_account import Credentials
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables
env_path = os.path.join(os.path.dirname(__file__), '..', 'script-dependencies', '.env')
load_dotenv(dotenv_path=env_path)

# Copy and paste the name of your Google Sheets API credentials file
credentials_filename = 'data-templates-c7159dc891a7.json' # EDIT ME
credentials_path = os.path.join(os.path.dirname(os.path.realpath(env_path)), credentials_filename)

# Copy and paste MIMARKS package Excel filename
mimarks_filename = 'MIMARKS.survey.sediment.6.0.xlsx' # EDIT ME
mimarks_file_path = os.path.join(os.path.dirname(os.path.realpath(env_path)), mimarks_filename)

# Get Google Sheet IDs from .env file
#study_template_dict_sheet_id = os.getenv("STUDY_TEMPLATE_DICT_SHEET_ID")
#new_template_id = os.getenv("NEW_TEMPLATE_ID")

#Alternatively, you can directly insert the Google Sheet IDs here instead:
study_template_dict_sheet_id = "1u4o-Cre-t5NAlwYoa6f-XTIDj8WrH05GK4ojS1y2FQA" #EDIT ME
new_template_id = "1auJth8xuGYVLAGik5QVEGSnSpGIYlIkdZLp_Bd20D64" #EDIT ME

# Define scopes for Google API
SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

# Authenticate with Google using the credentials file
creds = Credentials.from_service_account_file(credentials_path, scopes=SCOPES)
client = gspread.authorize(creds)

def get_mimarks_terms(mimarks_file_path):
    """Function to get all terms and their comments from the latest MIMARKs package"""
    print('\n----------------------------------------------')
    print('Getting MIMARKS terms from latest package...')
    
	# Load the MIMARKS Excel file using openpyxl to get comments
    wb = load_workbook(mimarks_file_path, data_only=True)
    ws = wb.active

    # Dictionary with term name as keys and comments as values
    mimarks_terms_with_comments = {}

    print("\nEnter the row number that contains the term names.")
    choice = int(input("Enter row number. Most MIMARKs packages have terms in row 12, or sometimes 13: "))

    # Iterate through the columns to get comments
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=choice, column=col)
        term = cell.value
        comment = cell.comment.text if cell.comment else None
        mimarks_terms_with_comments[term] = comment

    # Print the dictionary to verify
    print('\n------------MIMARKS package terms--------------')
    print('Printing MIMARKs Terms with their Comments')
    print('\n')
    for term, comment in mimarks_terms_with_comments.items():
        print(f"{term}:, {comment}")
    print('\n----------------------------------------------')
    
    return mimarks_terms_with_comments

def get_aoml_darwincore_terms(study_template_dict_sheet_id):
    """Function to get AOML and DarwinCore terms from the study-data-template-dict Google Sheet"""
    print('\n----------------------------------------------')
    print('\nLoading data dictionary...')
    
	#Preparing to read Google Sheet
    study_template_dict_sheet = client.open_by_key(study_template_dict_sheet_id)
    study_template_dict = study_template_dict_sheet.worksheet('study-data-templates')
    
    # Load the sheet data into a DataFrame, starting from row 2 to get the correct headers
    data = study_template_dict.get_all_values()
    
    print("\nGoogle Sheet loaded. Getting terms...")

    headers = data[1]  # Term names in row 2
    study_template_dict_df = pd.DataFrame(data[2:], columns=headers)  # Data starts from row 3

	# Regex to match all 'sheet' column values that CONTAIN sample_data
    pattern = r'\b(sample_data)\b'

    # Rows where 'sheet' column contains 'sample_data'
    filtered_data = study_template_dict_df[study_template_dict_df['sheet'].str.contains(pattern, na=False, regex=True)]

    # Grab all row 'name's (1st column) that were found above ^
    aoml_darwincore_terms = filtered_data.iloc[:, 0].tolist()

    print('\n------------AOML + DarwinCore Terms--------------')
    print(aoml_darwincore_terms)
    print('-------------------------------------------------')
    
    return aoml_darwincore_terms

def column_letter(col_idx):
    """Code utility function. Convert a column index into a column letter (1-indexed)."""
    string = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26) #26 for alphabet length
        string = chr(65 + remainder) + string #65 is ASCII For A
    return string

def edit_template(mimarks_terms_with_comments, new_template_id, study_template_dict_sheet_id):
    """Creates a new data template based off a previous one. """
    new_sheet = client.open_by_key(new_template_id)
    print('\n----------------------------------------------')
    data_template_type = input("Enter name of new data template type, i.e.: sediment, water, soil: ")
    sample_data_sheetname = data_template_type + '_sample_data'
    sample_data = new_sheet.worksheet(sample_data_sheetname)

	# Extra work done to move date_modified and modified_by columns to the end, after adding new terms
	# date_modified = dm, and modified_by = mb, in terms of the code
    existing_terms = sample_data.row_values(9)
    core_terms = existing_terms[:-2]
    dm_mb = existing_terms[-2:]
    
	# Temporarily remove '*' from term names for operations, but retain them for future use
    valid_terms = {term.replace('*', '').strip() for term in mimarks_terms_with_comments.keys()}
    valid_terms.update(get_aoml_darwincore_terms(study_template_dict_sheet_id))
    terms_to_delete = [term for term in core_terms if term not in valid_terms]
    terms_to_add = [term for term, comment in mimarks_terms_with_comments.items() if term.replace('*', '').strip() not in core_terms]
    
    print('\nTerms to add to data template: ')
    print(terms_to_add)
    print('\nTerms to be deleted from data template: ')
    print(terms_to_delete)
    
    new_MIMARKS_terms_with_comments = {term.replace('*', '').strip(): comment for term, comment in mimarks_terms_with_comments.items() if term.replace('*', '').strip() not in core_terms}
    print("\nNew terms to be added: ")
    for term in new_MIMARKS_terms_with_comments.items():
        print(f"{term}")
    
    for term in reversed(terms_to_delete):
        col_index = core_terms.index(term) + 1
        sample_data.delete_columns(col_index)
        core_terms.remove(term)

	# list_of_terms is used in the update_data_dictionary function to update new and old terms
    list_of_terms = core_terms + [term.replace('*', '').strip() for term in terms_to_add]
    
    updated_terms = core_terms + [term.replace('*', '').strip() for term in terms_to_add] + dm_mb
    # print("\nTotal terms after update (including date_modified and modified_by):", len(updated_terms))

    end_col_letter = column_letter(len(updated_terms))
    range_to_update = f'A9:{end_col_letter}9'
    # print("\nRange to update:", range_to_update)

    sample_data.update(range_name=range_to_update, values=[updated_terms])

	# Section below is to add associated comment to each term
    requests = []
    format_requests = []  # This will store formatting requests
    for i, term in enumerate(updated_terms, start=1):
        original_term = core_terms + terms_to_add + dm_mb
        cell = rowcol_to_a1(9, i)
        actual_term = original_term[i-1]
        if actual_term in terms_to_add:
            # Add comment to the cell
            comment = mimarks_terms_with_comments.get(actual_term, '')
            if comment:
                requests.append({
                    "updateCells": {
                        "range": {
                            "sheetId": sample_data.id,
                            "startRowIndex": 8,
                            "endRowIndex": 9,
                            "startColumnIndex": i-1,
                            "endColumnIndex": i
                        },
                        "rows": [{
                            "values": [{
                                "note": comment
                            }]
                        }],
                        "fields": "note"
                    }
                })

            # Determine the background color based on the presence of '*'
            if actual_term.startswith('*'):
                bgcolor = Color(0.0, 1.0, 0.0)  # Google sheets Green
            else:
                bgcolor = Color(1.0, 1.0, 0.0)  # Google Sheets yellow

            format_requests.append((cell, CellFormat(backgroundColor=bgcolor)))

    # Execute all requests in one batch update to avoid multiple API calls
    if requests:
        new_sheet.batch_update({'requests': requests})

    # Apply all formatting requests
    if format_requests:
        format_cell_ranges(sample_data, format_requests)

    print("\nGoogle Sheet data template editing complete.")
    print('----------------------------------------------')
    
    return(sample_data_sheetname, new_MIMARKS_terms_with_comments, list_of_terms)

def update_data_dictionary(sample_data_sheetname, new_MIMARKS_terms_with_comments, study_template_dict_sheet_id, list_of_terms):
    """Function to add new terms to data dictionary, and update the 'sheet' column for existing terms"""
    print('----------------------------------------------')
    print('Loading study-data-template-dict Google sheet...')
	# Establish connection to the Google Sheets document
    study_sheet = client.open_by_key(study_template_dict_sheet_id)
    study_template_dict = study_sheet.worksheet('study-data-templates')

    # Fetch AOML and DarwinCore terms to exclude
    aoml_darwincore_terms = get_aoml_darwincore_terms(study_template_dict_sheet_id)
    
    print('\nMaking updates to study-data-template-dict...')

    # Remove AOML/DarwinCore terms and new MIMARKS terms from the list_of_terms
    filtered_terms = [term for term in list_of_terms if term not in aoml_darwincore_terms and term not in new_MIMARKS_terms_with_comments.keys()]

    # Fetch current data to DataFrame for easier manipulation
    data = study_template_dict.get_all_values()
    if data:  # Check if data is not empty
        headers = data[1]  # Assuming the second row contains headers
        df = pd.DataFrame(data[2:], columns=headers)  # Data starts from the third row

        # Trim whitespace from column headers to avoid issues
        df.columns = df.columns.str.strip()

        # Update 'sheet' column for terms that need updating
        updated = False
        for index, row in df.iterrows():
            if row['name'] in filtered_terms and sample_data_sheetname not in row['sheet'].split('|'):
                updated_sheets = row['sheet'] + ' | ' + sample_data_sheetname
                df.at[index, 'sheet'] = updated_sheets
                updated = True

        # Convert the DataFrame back to a list of lists to update the Google Sheet if updates were made
        if updated:
            updated_data = df.values.tolist()
            study_template_dict.update('A3', updated_data, value_input_option='USER_ENTERED')

        # Get the last row number to start appending new rows
        last_row = len(df) + 3  # Adding 3 because data starts from row 3

        # Add new MIMARKS terms
        for term, comment in new_MIMARKS_terms_with_comments.items():
            # Prepare the row content based on the column structure
            row_content = [term, '', '', comment, '', sample_data_sheetname]
            # Append the row to the sheet
            study_template_dict.append_row(row_content, table_range=f"A{last_row}")

            last_row += 1  # Increment to the next row for the next term

    else:
        print("Error, no datasheet found.")

    print('\nUpdate complete. Specified terms have been updated and new terms added to the dictionary.')
    print('----------------------------------------------')

def add_new_sheet_to_dict(sample_data_sheetname, list_of_terms, study_template_dict_sheet_id):
    try:
        # Establish connection to the Google Sheets document
        study_sheet = client.open_by_key(study_template_dict_sheet_id)
        sample_data_sheet = study_sheet.worksheet(sample_data_sheetname)
        study_data_templates_sheet = study_sheet.worksheet('study-data-templates')

        # Clear existing data to avoid duplicates
        sample_data_sheet.clear()

        # Set the headers with bold formatting
        headers = ["Term", "Definition"]
        sample_data_sheet.append_row(headers)
        gf.format_cell_range(sample_data_sheet, 'A1:B1', gf.CellFormat(textFormat=gf.TextFormat(bold=True)))

        # Fetch all data starting from row 3 (data section starts here)
        raw_data = study_data_templates_sheet.get_all_values()[2:]  # Data starts from row 3

        # Prepare the data to be inserted into the new sheet
        data_to_insert = []
        for term in list_of_terms:
            for row in raw_data:
                if row[0].strip() == term:  # Match term exactly in column A
                    term_link = f"[{term}](https://noaa-omics-templates.readthedocs.io/en/latest/terms/sample_data/{term}.html)"
                    definition = row[3] if len(row) > 3 else ''  # Get definition from column D, handle missing data
                    data_to_insert.append([term_link, definition])
                    break

        # Determine the range to update
        start_row = 2  # Start from the second row because the first row is headers
        end_row = start_row + len(data_to_insert) - 1
        range_to_update = f'A{start_row}:B{end_row}'

        # Batch update to add all rows at once
        sample_data_sheet.update(range_to_update, data_to_insert)

        print('Update complete. New terms and definitions added to the sample data sheet.')

    except Exception as e:
        print(f"An error occurred while updating the new sheet: {e}")


def get_color_values_of_row(new_template_id, sample_data_sheetname):
    """
    Reads the background colors for terms in a specified row from the data template Google Sheet,
    processing 50 terms at a time to stay within API rate limits.
    """
    data_template_sheet = client.open_by_key(new_template_id).worksheet(sample_data_sheetname)
    term_row = int(input("Enter the row number where terms are located with colors in the data template: "))
    terms = data_template_sheet.row_values(term_row)
    
    term_colors = {}
    total_terms = len(terms)
    chunk_size = 50  # Number of terms to process in each batch (avoiding API read request limits)

    omit_last_two = input("Would you like to omit the last 2 columns (typically date_modified and modified_by, for internal use (Y / N): ")
    if omit_last_two.lower() == 'yes':
        total_terms -= 2 #reduce count by 2 to omit last 2 columns

    # Process each chunk
    for start in range(0, total_terms, chunk_size):
        end = start + chunk_size
        chunk_terms = terms[start:end]
        for i, term in enumerate(chunk_terms, start=start):
            col_label = column_letter(i + 1)
            cell_address = f"{col_label}{term_row}"
            format_info = get_effective_format(data_template_sheet, cell_address)
            bg_color = format_info.backgroundColor
            
            red = bg_color.red if bg_color.red is not None else 0
            green = bg_color.green if bg_color.green is not None else 0
            blue = bg_color.blue if bg_color.blue is not None else 0
            color_key = f"{red:.2f}, {green:.2f}, {blue:.2f}"
            
            term_colors[term] = color_key
            print(f"Term '{term}' at column {col_label} has color {color_key}")

        # Wait if there are still more terms to process
        # You will get an API Read request error for too many reads if you remove the sleep time
        if end < total_terms:
            print("Waiting 60 seconds before processing the next batch...")
            time.sleep(60)  # sometimes works with as little as 35 seconds wait time...but the 'sweet spot' seems to be a minute

    

    print("------------------------------------")
    print("Printing term_colors: ")
    print(term_colors)
    return term_colors

def update_dropdown_values_from_colors(sample_data_sheetname, study_template_dict_sheet_id, term_colors):
    """
    Updates dropdown values in the study-data-template's new term sheet based on the term's colors in the data template
    """
    # Open the worksheet
    study_template_dict_sheet = client.open_by_key(study_template_dict_sheet_id).worksheet(sample_data_sheetname)

    # Define the color to requirement mapping
    color_to_requirement = {
        "0.57, 0.82, 0.31": "NCBI+OBIS",  # Custom green
        "0.00, 1.00, 0.00": "NCBI+OBIS",  # Default green
        "1.00, 0.60, 0.00": "Recommended",  # Default Orange
        "1.00, 1.00, 0.00": "Optional"      # Default Yellow
    }

    # Insert the header in the first row, column C
    study_template_dict_sheet.update_acell('C1', 'required_by')  # Set header directly

    # Fetch all rows starting from row 2 (data starts on row 2)
    rows = study_template_dict_sheet.get_all_values()[1:]  # Exclude header row

    # Prepare updates for batch processing
    updates = []

    # Iterate over each row, extract term names, and update the dropdown
    for i, row in enumerate(rows, start=2):  # Start at 2 because data starts from the second row
        term_url = row[0]
        term_name = term_url[term_url.find('[')+1:term_url.find(']')]  # Extract the term name from brackets

        # Check if the term name is in the term_colors dictionary
        if term_name in term_colors:
            rgb_code = format_rgb_code(term_colors[term_name])  # Ensure RGB codes are formatted to match keys
            dropdown_value = color_to_requirement.get(rgb_code, "Check Color")  # Map RGB to dropdown value or default
            
            # Prepare the update for this row's dropdown cell
            updates.append({
                "range": f"C{i}",  # Correct indexing for 1-based API and aligns directly with row numbers
                "values": [[dropdown_value]]
            })

    # Batch update the cells if there are updates to be made
    if updates:
        study_template_dict_sheet.batch_update(updates)

    print(f"Updated {len(updates)} dropdown values based on term colors.")

def format_rgb_code(rgb_code):
    """
    Sanitizes RGB values received from API (truncates after 2nd decimal place for long floats) 
    """
    # Split and format each component to two decimal places
    r, g, b = map(float, rgb_code.split(', '))
    return f"{r:.2f}, {g:.2f}, {b:.2f}"

def verify_data_dictionary(study_template_dict_sheet_id):
    print('----------------------------------------------')
    print('Verifying study data template dictionary after edits...')

    # Establish connection to the Google Sheets document
    study_sheet = client.open_by_key(study_template_dict_sheet_id)
    study_data_templates_sheet = study_sheet.worksheet('study-data-templates')

    # Fetch current data to DataFrame for easier manipulation
    data = study_data_templates_sheet.get_all_values()
    if len(data) < 3:
        print("Insufficient data in the sheet.")
        return

    headers = [header.strip() for header in data[1]]  # Strip spaces from headers
    df = pd.DataFrame(data[2:], columns=headers)  # Data starts from row 3

    # Find duplicates based on the 'name' column
    duplicate_mask = df.duplicated('name', keep=False)
    duplicates = df[duplicate_mask]

    if duplicates.empty:
        print("No duplicates found.")
    else:
        print(f"Found duplicates: {duplicates['name'].tolist()}")

        # Group by name to handle each set of duplicates
        for name, group in duplicates.groupby('name'):
            print(f"\nHandling duplicates for term: {name}")
            rows = group.index.tolist()
            old_idx = min(rows)  # The old entry
            new_idx = max(rows)  # The new entry (higher row number)

            # Show user duplicates and ask for action
            print(f"Old entry 'sheet': {df.at[old_idx, 'sheet']}")
            print(f"New entry 'sheet': {df.at[new_idx, 'sheet']}")
            action = input("Do you want to merge these entries? (yes/no): ")

            if action.lower() == 'yes':
                # Check if 'sheet' values are different
                if df.at[new_idx, 'sheet'] not in df.at[old_idx, 'sheet']:
                    # Append new 'sheet' value to the old entry
                    updated_sheets = df.at[old_idx, 'sheet'] + ' | ' + df.at[new_idx, 'sheet']
                    df.at[old_idx, 'sheet'] = updated_sheets
                    print(f"Updated 'sheet' for {name} at row {old_idx + 3}: {updated_sheets}")

                    # Update the sheet to reflect changes
                    updated_data = df.values.tolist()
                    study_data_templates_sheet.update('A3', updated_data, value_input_option='USER_ENTERED')

                # Delete the new duplicate row after confirming merge
                study_data_templates_sheet.delete_rows(new_idx + 3)  # Adjusting for header rows
                print(f"Deleted newest duplicate for {name} at row {new_idx + 3}")

    print('----------------------------------------------')


if __name__ == "__main__":
    # Retrieve MIMARKS terms with comments
    mimarks_terms_with_comments = get_mimarks_terms(mimarks_file_path)
    sample_data_sheetname, new_MIMARKS_terms_with_comments, list_of_terms = edit_template(mimarks_terms_with_comments, new_template_id, study_template_dict_sheet_id)
    
    print('Please check your new data template for accuracy before continuing!')
    print('If your data template is correct, update study-data-template-dict to include the new terms')
    
    # First check for continuation
    choice = input('Would you like to continue and update the data dictionary? (Y / N): ').lower()
    if choice in ['y', 'yes']:
        update_data_dictionary(sample_data_sheetname, new_MIMARKS_terms_with_comments, study_template_dict_sheet_id, list_of_terms)
        verify_data_dictionary(study_template_dict_sheet_id)
    else:
        print('Exiting program. Restore your data template to the previous version in Google Sheets if you discovered errors.')
        sys.exit()
    
    print(f"If everything is correct to this point, please duplicate an existing <environment>_sample_data table in the data dictionary, and rename it to: '{sample_data_sheetname}'.")
    choice2 = input('Have you created the new sheet and would you like to continue? (Y / N): ').lower()
    if choice2 in ['y', 'yes']:
        add_new_sheet_to_dict(sample_data_sheetname, list_of_terms, study_template_dict_sheet_id)
    else:
        print('Exiting program. Restore the Google Sheets to their previous completed version.')
        sys.exit()
    
    # Getting term colors and updating dropdown values
    print("Proceeding to update dropdown values based on term colors.")
    term_colors = get_color_values_of_row(new_template_id, sample_data_sheetname)
    update_dropdown_values_from_colors(sample_data_sheetname, study_template_dict_sheet_id, term_colors)
    print("Dropdown values updated successfully.")