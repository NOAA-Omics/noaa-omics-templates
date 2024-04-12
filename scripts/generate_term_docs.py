import sys
import os
from openpyxl import load_workbook

def read_excel_file(file_path):
    wb = load_workbook(filename=file_path, read_only=False)
    ws = wb.active

    data = {}
    #skips the 1st 2 rows
    r= 3

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:
            key = row[0]
            hyperlink_text = row[4]
            hyperlink_target = None
            try:
            # Check if the cell contains a hyperlink
                if ws.cell(row=r, column=5).hyperlink:
                    hyperlink_target = ws.cell(row=r, column=5).hyperlink.target
                else: hyperlink_target = None
            except AttributeError:
                hyperlink_target = None

            data[key] = {
                'example': row[1],
                'format': row[2],
                'definition': row[3],
                'sheet': row[5],
                'required_by': row[6],
                'origin_text': hyperlink_text,
                'origin_link': hyperlink_target
            }
        r += 1

    wb.close()
    return data

def write_mds(data, folder_name):
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
    for key, value in data.items():
        sheets_dict = {}
        if '|' in value['sheet']:
            sheets = value['sheet'].split(" | ")
            for s in sheets:
                sheets_dict[s] = os.path.join(folder_name,s)
        else:
            sheets_dict[value['sheet']] = os.path.join(folder_name, f"{value['sheet']}")
        for s in sheets_dict.keys():
            if not os.path.exists(sheets_dict[s]):
                os.makedirs(sheets_dict[s])
            file_name = os.path.join(sheets_dict[s], f"{key}.md")
            with open(file_name, 'w') as file:
                file.write(f"# Term: {key}\n\n")
                file.write(f"*{value['definition']}*\n\n")
                if value['origin_link'] == None:
                    file.write(f"Origin: {value['origin_text']}\n\n")
                else:
                    file.write(f"Origin: [{value['origin_text']}]({value['origin_link']})\n\n")
                file.write(f"Example: {value['example']}\n\n")
                file.write(f"Sheet(s) containing term: {value['sheet']}\n\n")
                if value['required_by'] == 'NCBI+OBIS':
                    file.write("**Required by NCBI and OBIS**")
                elif value['required_by'] == 'Optional':
                    file.write("**Optional or context dependent**")
                elif value['required_by'] == 'Recommended':
                    file.write("**Not required, but recommended by NOAA Omics**")
                elif value['required_by'] == 'NCBI':
                    file.write("**Required by NCBI**")
                elif value['required_by'] == 'OBIS':
                    file.write("**Required by OBIS**")



if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_term_docs.py <file_path> <output_folder>")
    else:
        file_path = sys.argv[1]
        folder_name = sys.argv[2]
        data_dict = read_excel_file(file_path)
        write_mds(data_dict, folder_name)
        