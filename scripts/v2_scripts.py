import os
import re
import time
import sys
import gspread
import gspread_formatting as gf
from gspread_formatting import *
from gspread.utils import rowcol_to_a1
import pandas as pd
from google.oauth2.service_account import Credentials
from dotenv import load_dotenv
from openpyxl import load_workbook

#--------------------------------- FILE INPUT AND CREDENTIALS ---------------------------------#
# Load environment variables
env_path = os.path.join(os.path.dirname(__file__), '..', 'script-dependencies', '.env')
load_dotenv(dotenv_path=env_path)

# Copy and paste the name of your Google Sheets API credentials file
credentials_filename = 'data-templates-c7159dc891a7.json'  # EDIT ME
credentials_path = os.path.join(os.path.dirname(os.path.realpath(env_path)), credentials_filename)

# Copy and paste Excel filename to read or edit
excel_filename = 'Events_output.xlsx'  # EDIT ME
excel_file_path = os.path.join(os.path.dirname(os.path.realpath(env_path)), excel_filename)

if os.path.exists(excel_file_path):
    print(f"Excel file, {excel_filename}, imported successfully.")
else:
    print(f"Excel file {excel_filename} not found. If you're not using an Excel file, please ignore this.")
    print("HINT: Ensure your Excel filename is correct (with the .xlsx extension), and make sure your Excel file is in the 'script-dependencies' folder.")

# Define Google Sheet IDs
google_sheet_id_1 = "1u4o-Cre-t5NAlwYoa6f-XTIDj8WrH05GK4ojS1y2FQA"  # EDIT ME
google_sheet_id_2 = "1uwD6zUMCAlUn_xqbTufKeTUXL5F_n8rH1NzKBUKL3ks"  # EDIT ME
# This id_2 is for the NEW data template COPY in Automation Testing

# Define scopes for Google API
SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]

# Authenticate with Google using the credentials file
creds = Credentials.from_service_account_file(credentials_path, scopes=SCOPES)
client = gspread.authorize(creds)

# Initialize variables to retain values between loop iterations
google_sheet_id_1_terms = None
google_sheet_id_2_terms = None
mimarks_terms_with_comments = None
sample_data_sheetname = None
new_MIMARKS_terms_with_comments = None
list_of_terms = None
term_colors = None

#--------------------------------- FUNCTIONS ---------------------------------#

def get_terms_row_GS(google_sheet_id, sheet_name, row_num, as_dict=False):
    sheet = client.open_by_key(google_sheet_id).worksheet(sheet_name)
    terms = sheet.row_values(row_num)
    if as_dict:
        data = sheet.row_values(row_num + 1)
        terms_dict = {term: data[idx] if idx < len(data) else None for idx, term in enumerate(terms)}
        return terms_dict
    return terms

def get_terms_row_X(excel_filename, row_num, as_dict=False):
    wb = load_workbook(excel_filename, data_only=True)
    ws = wb.active
    terms = [ws.cell(row=row_num, column=col).value for col in range(1, ws.max_column + 1)]
    if as_dict:
        data = [ws.cell(row=row_num + 1, column=col).value for col in range(1, ws.max_column + 1)]
        terms_dict = {term: data[idx] if idx < len(data) else None for idx, term in enumerate(terms)}
        return terms_dict
    return terms

def get_terms_column_GS(google_sheet_id, sheet_name, col_num):
    sheet = client.open_by_key(google_sheet_id).worksheet(sheet_name)
    terms = sheet.col_values(col_num)
    return terms

def get_terms_col_X(excel_filename, col_num):
    wb = load_workbook(excel_filename, data_only=True)
    ws = wb.active
    terms = [ws.cell(row=row, column=col_num).value for row in range(1, ws.max_row + 1)]
    return terms

def term_merger():
    # Placeholder for term merging function
    print("Term merger function not implemented yet.")

def print_collected_data():
    print("\nCollected Data:")
    print(f"Google Sheet 1 Terms: {google_sheet_id_1_terms}")
    print(f"Google Sheet 2 Terms: {google_sheet_id_2_terms}")
    print(f"MIMARKS Terms with Comments: {mimarks_terms_with_comments}")
    print(f"Sample Data Sheet Name: {sample_data_sheetname}")
    print(f"New MIMARKS Terms with Comments: {new_MIMARKS_terms_with_comments}")
    print(f"List of Terms: {list_of_terms}")
    print(f"Term Colors: {term_colors}")
    print("----------------------------------------------\n")

def main():
    global google_sheet_id_1_terms, google_sheet_id_2_terms, mimarks_terms_with_comments
    global sample_data_sheetname, new_MIMARKS_terms_with_comments, list_of_terms, term_colors

    print('Please be sure to set up your files and credentials before trying to run a function.')
    while True:
        print('\nPlease select a function to run. NOTE: If you are running multiple functions, do not exit the program.')
        print('Your lists, dictionaries, files, and credentials are retained between runs of this program.')
        print("'1': Get terms/column names from Google Sheets")
        print("'2': Get all terms in a column from Google Sheets")
        print("'3': Get term/column names from Excel file")
        print("'4': Get all terms in a column from Excel file")
        print("'5': Term merger")
        print("'print' or 'p': Print collected data")
        print("'0': Exit")

        user_input = input('Enter the corresponding number of the function you want to run: ').strip()

        if user_input == '1':
            google_sheet_id = int(input("To read the first Google Sheet, enter '1', for the 2nd one, enter '2': "))
            sheet_name = input('Enter the sheet name: ').strip()
            row_num = int(input('Enter the row number of the terms: '))
            as_dict = input('Do you want the terms as a dictionary with the first piece of data? (yes/no): ').strip().lower() == 'yes' | 'y'
            if google_sheet_id == 1:
                google_sheet_id_1_terms = get_terms_row_GS(google_sheet_id_1, sheet_name, row_num, as_dict=as_dict)
                print(f"\nTerms from Google Sheet {google_sheet_id_1} (Sheet: {sheet_name}, Row: {row_num}):")
                print(google_sheet_id_1_terms)
            elif google_sheet_id == 2:
                google_sheet_id_2_terms = get_terms_row_GS(google_sheet_id_2, sheet_name, row_num, as_dict=as_dict)
                print(f"\nTerms from Google Sheet {google_sheet_id_2} (Sheet: {sheet_name}, Row: {row_num}):")
                print(google_sheet_id_2_terms)
            else:
                print('Invalid input. Try again and enter either 1 or 2.')

        elif user_input == '2':
            google_sheet_id = int(input("To read the first Google Sheet, enter '1', for the 2nd one, enter '2': "))
            sheet_name = input('Enter the sheet name: ').strip()
            col_num = int(input('Enter the column number of the terms: '))
            if google_sheet_id == 1:
                google_sheet_id_1_terms = get_terms_column_GS(google_sheet_id_1, sheet_name, col_num)
                print(f"\nTerms from Google Sheet {google_sheet_id_1} (Sheet: {sheet_name}, Column: {col_num}):")
                print(google_sheet_id_1_terms)
            elif google_sheet_id == 2:
                google_sheet_id_2_terms = get_terms_column_GS(google_sheet_id_2, sheet_name, col_num)
                print(f"\nTerms from Google Sheet {google_sheet_id_2} (Sheet: {sheet_name}, Column: {col_num}):")
                print(google_sheet_id_2_terms)
            else:
                print('Invalid input. Try again and enter either 1 or 2.')

        elif user_input == '3':
            row_num = int(input('Enter the row number of the terms in the Excel file: '))
            as_dict = input('Do you want the terms as a dictionary with the first piece of data? (yes/no): ').strip().lower() == 'yes' | 'y'
            mimarks_terms_with_comments = get_terms_row_X(excel_file_path, row_num, as_dict=as_dict)
            print(f"\nTerms from Excel file {excel_filename} (Row: {row_num}):")
            print(mimarks_terms_with_comments)

        elif user_input == '4':
            col_num = int(input('Enter the column number of the terms in the Excel file: '))
            mimarks_terms_with_comments = get_terms_col_X(excel_file_path, col_num)
            print(f"\nTerms from Excel file {excel_filename} (Column: {col_num}):")
            print(mimarks_terms_with_comments)

        elif user_input == '5':
            term_merger()

        elif user_input == 'print' or user_input == 'p':
            print_collected_data()

        elif user_input == '0':
            print("Exiting program.")
            break

        else:
            print('Invalid input. Try again.')

if __name__ == "__main__":
    main()
