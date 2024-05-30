#This script reads the new terms from the MIMARKS package and inserts them into the new (copied) template
#This script will probably be removed. It modifies Excel files. No Google sheets interaction

from openpyxl import load_workbook

def insert_new_terms(template_path, mimarks_path, sheet_name):
    # Load the template and MIMARKS files
    template_wb = load_workbook(template_path)
    template_ws = template_wb[sheet_name]

    mimarks_wb = load_workbook(mimarks_path)
    mimarks_ws = mimarks_wb.active

    # Collect existing terms in the template
    template_terms = {row[0].value for row in template_ws.iter_rows(min_row=2, values_only=True)}
    new_terms = []

    # Identify new terms from the MIMARKS file
    for row in mimarks_ws.iter_rows(min_row=2, values_only=True):
        if row[0] not in template_terms:
            new_terms.append(row)

    # Append new terms to the template
    for term in new_terms:
        template_ws.append(term)

    # Save the updated template
    template_wb.save(template_path)

if __name__ == "__main__":
    template_path = ''
    mimarks_path = ''
    sheet_name = ''
    insert_new_terms(template_path, mimarks_path, sheet_name)
